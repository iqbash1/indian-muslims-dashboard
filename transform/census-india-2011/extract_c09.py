"""
L1 -> L2 for Census 2011 C-9 Educational Level by Religious Community.

Reads:  sources/census-2011/c-series/c09-education-by-religion.xlsx
Writes: extracted/census-2011/c09-education-by-religion.csv

Slim version: extracts only the Total Population, Illiterate, and Literate
columns (cols 6-14). The educational-level breakdowns (cols 15-41) are not
extracted here — add a separate extractor when a metric needs them.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import pathlib
import sys

from openpyxl import load_workbook

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
SOURCE_PATH = REPO_ROOT / "sources" / "census-2011" / "c-series" / "c09-education-by-religion.xlsx"
OUTPUT_PATH = REPO_ROOT / "extracted" / "census-2011" / "c09-education-by-religion.csv"
EXTRACTOR_VERSION = "1.0.0"

# Column layout per the C-09 sheet header (rows 2-5).
# Each measure occupies 3 consecutive columns: persons, males, females.
MEASURE_COL_GROUPS = [
    ("total_population", 6),
    ("illiterate", 9),
    ("literate", 12),
]
SEX_LABELS = ["persons", "males", "females"]
RESIDENCE_MAP = {"Total": "total", "Rural": "rural", "Urban": "urban"}
RELIGION_MAP = {
    "All religious communities": "all",
    "Hindu": "hindu",
    "Muslim": "muslim",
    "Christian": "christian",
    "Sikh": "sikh",
    "Buddhist": "buddhist",
    "Jain": "jain",
    "Other religions and persuasions": "other",
}


def sha256_of(path: pathlib.Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def verify_source_integrity() -> dict:
    meta_path = SOURCE_PATH.with_suffix(SOURCE_PATH.suffix + ".meta.json")
    meta = json.loads(meta_path.read_text())
    actual_sha = sha256_of(SOURCE_PATH)
    if actual_sha != meta["sha256"]:
        sys.exit(
            f"sha256 mismatch for {SOURCE_PATH.name}: "
            f"archive {actual_sha[:16]} != sidecar {meta['sha256'][:16]}"
        )
    return meta


def extract() -> None:
    meta = verify_source_integrity()
    wb = load_workbook(str(SOURCE_PATH), read_only=True)
    ws = wb["C-09"]

    extraction_run = (
        f"census-c09-extract-v{EXTRACTOR_VERSION}-"
        f"{dt.datetime.now(dt.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    n_rows = 0
    n_skipped_religion = 0
    n_skipped_residence = 0

    with OUTPUT_PATH.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "source_id", "source_document", "source_sha256_prefix", "extraction_run",
            "table_name", "state_code", "area_name", "residence", "religion",
            "age_group", "measure", "sex", "value",
        ])

        # Data starts at row 8 (rows 1-7 are headers + blank).
        for row in ws.iter_rows(min_row=8, values_only=True):
            area = row[3]
            religion_raw = row[4]
            tru_raw = row[2]
            age = row[5]
            if not area or not religion_raw or not tru_raw or age in (None, ""):
                continue

            residence = RESIDENCE_MAP.get(str(tru_raw).strip())
            if residence is None:
                n_skipped_residence += 1
                continue
            religion = RELIGION_MAP.get(str(religion_raw).strip())
            if religion is None:
                n_skipped_religion += 1
                continue

            table = str(row[0] or "").strip()
            state = str(row[1] or "").strip()
            area_clean = str(area).strip()
            age_str = str(age).strip()

            for measure, base_col in MEASURE_COL_GROUPS:
                for sex_offset, sex in enumerate(SEX_LABELS):
                    raw = row[base_col + sex_offset]
                    if raw in (None, ""):
                        continue
                    value = int(raw) if isinstance(raw, (int, float)) else int(float(raw))
                    w.writerow([
                        "census-india-2011",
                        str(SOURCE_PATH.relative_to(REPO_ROOT)),
                        meta["sha256"][:16],
                        extraction_run,
                        table, state, area_clean, residence, religion,
                        age_str, measure, sex, value,
                    ])
                    n_rows += 1

    print(
        f"wrote {OUTPUT_PATH.relative_to(REPO_ROOT)} "
        f"({n_rows} rows; skipped {n_skipped_religion} unknown-religion, "
        f"{n_skipped_residence} unknown-residence)"
    )


if __name__ == "__main__":
    extract()
