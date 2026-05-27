"""
L1 -> L2 for Census 2011 C-15 Religious Community by Age-Group and Sex.

Reads:  sources/census-2011/c-series/c15-religion-by-age-sex.xlsx
Writes: extracted/census-2011/c15-religion-by-age-sex.csv

Long-format: one row per (area x residence x age_group x religion x sex).
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import pathlib
import sys

from openpyxl import load_workbook

REPO_ROOT = pathlib.Path(__file__).resolve().parents[2]
SOURCE_PATH = REPO_ROOT / "sources" / "census-2011" / "c-series" / "c15-religion-by-age-sex.xlsx"
OUTPUT_PATH = REPO_ROOT / "extracted" / "census-2011" / "c15-religion-by-age-sex.csv"
EXTRACTOR_VERSION = "1.0.0"

# Column layout per the C-15 sheet header (rows 2-4).
RELIGION_COL_GROUPS = [
    ("all", 6),
    ("hindu", 9),
    ("muslim", 12),
    ("christian", 15),
    ("sikh", 18),
    ("buddhist", 21),
    ("jain", 24),
    ("other", 27),
]
SEX_LABELS = ["persons", "males", "females"]
RESIDENCE_MAP = {"Total": "total", "Rural": "rural", "Urban": "urban"}


def sha256_of(path: pathlib.Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def verify_source_integrity() -> dict:
    meta_path = SOURCE_PATH.with_suffix(SOURCE_PATH.suffix + ".meta.json")
    meta = json.loads(meta_path.read_text())
    actual_sha = sha256_of(SOURCE_PATH)
    if actual_sha != meta["sha256"]:
        sys.exit(
            f"sha256 mismatch for {SOURCE_PATH.name}: "
            f"archive {actual_sha[:16]} != sidecar {meta['sha256'][:16]}"
        )
    return meta


def extract() -> None:
    meta = verify_source_integrity()
    wb = load_workbook(str(SOURCE_PATH), read_only=True)
    ws = wb["C-15"]

    extraction_run = (
        f"census-c15-extract-v{EXTRACTOR_VERSION}-"
        f"{dt.datetime.now(dt.timezone.utc).strftime('%Y%m%dT%H%M%SZ')}"
    )

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    n_rows = 0
    with OUTPUT_PATH.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "source_id", "source_document", "source_sha256_prefix", "extraction_run",
            "table_name", "state_code", "distt_code", "area_name", "residence",
            "age_group", "religion", "sex", "value",
        ])

        # Data starts at row 8 (rows 1-7 are headers + blank).
        for row in ws.iter_rows(min_row=8, values_only=True):
            area = row[4]
            tru = row[3]
            age = row[5]
            if not area or not tru or age in (None, ""):
                continue
            residence = RESIDENCE_MAP.get(str(tru).strip())
            if residence is None:
                continue

            table = str(row[0] or "").strip()
            state = str(row[1] or "").strip()
            distt = str(row[2] or "").strip()
            area_clean = str(area).strip()
            age_str = str(age).strip()

            for religion, base_col in RELIGION_COL_GROUPS:
                for sex_offset, sex in enumerate(SEX_LABELS):
                    raw = row[base_col + sex_offset]
                    if raw in (None, ""):
                        continue
                    value = int(raw) if isinstance(raw, (int, float)) else int(float(raw))
                    w.writerow([
                        "census-india-2011",
                        str(SOURCE_PATH.relative_to(REPO_ROOT)),
                        meta["sha256"][:16],
                        extraction_run,
                        table, state, distt, area_clean, residence,
                        age_str, religion, sex, value,
                    ])
                    n_rows += 1

    print(f"wrote {OUTPUT_PATH.relative_to(REPO_ROOT)} ({n_rows} rows)")


if __name__ == "__main__":
    extract()
